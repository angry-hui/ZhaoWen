import pyodbc,os,sys,datetime,openpyxl
from PySide2.QtWidgets import QApplication,QMainWindow,QMessageBox
from Main_ui import Ui_mainWindow
import win32com.client
from PySide2.QtCore import QDate
from PySide2.QtGui import QStandardItem,QStandardItemModel
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side

cur_year = datetime.datetime.now().year
cur_month = datetime.datetime.now().month
cur_day = datetime.datetime.now().day
cur_date = str(cur_year) + "-" + ("00" + str(cur_month))[-2:] + "-" +("00" + str(cur_day))[-2:]
class Food_Analysis(Ui_mainWindow,QMainWindow):
    def __init__(self,parent=None):
        super(Food_Analysis,self).__init__(parent=None)
        self.setupUi(self)
        self.conn_db()
        self.Button_Catch.clicked.connect(self.catch_data)
        self.ComboBox_IP.addItems(['192.168.1.161','192.168.1.162','192.168.1.166','192.168.1.175','192.168.1.187'])
        self.DateEdit_Start.setDate(QDate(cur_year,cur_month,1))
        self.DateEdit_End.setDate(QDate(cur_year,cur_month,1))
        self.Button_Query.clicked.connect(self.Data_Query)
        self.Button_Ana.clicked.connect(self.Data_Ana)
        self.Button_Export.clicked.connect(self.Export_Data)
        self.Button_Clear.clicked.connect(self.Clear_mac)

    def conn_db(self):
        server = '192.168.1.254'
        database = 'Data_Analysis'
        username = 'sa'
        password = 'zhaowenerp.369'
        self.cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        self.cursor = self.cnxn.cursor()
    
    #清除卡机数据
    def Clear_mac(self):
        #kq_ip = ['192.168.1.161','192.168.1.162','192.168.1.166','192.168.1.175','192.168.1.187']
        zk = win32com.client.Dispatch('zkemkeeper.ZKEM.1')
        kq_ip = self.ComboBox_IP.currentText()
        if not zk.Connect_Net(kq_ip,4370):
            QMessageBox.information(self,'提示信息', "考勤机" + kq_ip + '连接连接失败!', QMessageBox.Yes)
            return    
        if zk.ClearGLog(1):
            QMessageBox.information(self,"提示信息","卡机:" + kq_ip + '清除成功!',QMessageBox.Yes)

            
    def catch_data(self):
        #kq_ip = ['192.168.1.161','192.168.1.162','192.168.1.166','192.168.1.175','192.168.1.187']
        kq_ip = ['192.168.1.187']
        for ip in kq_ip:
            zk = win32com.client.Dispatch('zkemkeeper.ZKEM.1')
            if not zk.Connect_Net(ip,4370):
                QMessageBox.information(self,'提示信息', "考勤机" + ip + '连接连接失败!',QMessageBox.Yes)
                return
            #读取所有的数据到内存中
            if zk.ReadAllGLogData(1):
                while 1:
                    dwMachineNumber, dwEnrollNumber, dwVerifyMode, dwInOutMode, dwYear, dwMonth, dwDay, dwHour, dwMinute, dwSecond, dwWorkcode = zk.SSR_GetGeneralLogData(1)
                    if not dwMachineNumber:
                        break
                    #print(dwEnrollNumber,dwYear,dwMonth,dwDay,dwHour,dwSecond,dwMinute)
                    kq_id = dwEnrollNumber
                    kq_date = str(dwYear) + "-" +  ("00" + str(dwMonth))[-2:] + "-" + ("00" + str(dwDay))[-2:]
                    kq_time = ("00" + str(dwHour))[-2:] + ":" + ("00" + str(dwSecond))[-2:]
                    params = ([kq_id])
                    sql = "select employeeid from zwerp.dbo.employee where iccardnumber = ?"
                    result = self.cursor.execute(sql,params).fetchone()           
                    if result is None or len(result) == 0:
                        continue
                    kq_employee = result[0]
                    params = ([kq_employee,kq_date,kq_time])
                    sql_counter = "select count(employeeid) from em_card_food where employeeid = ? and card_d = ? and card_tm = ?"
                    result = self.cursor.execute(sql_counter,params).fetchone()[0]
                    if result is None or result == 0 :
                        params = ([kq_employee,kq_date,kq_time,ip])
                        sql = "insert into em_card_food(employeeid,card_d,card_tm,card_ip) values (?,?,?,?)"
                        self.cursor.execute(sql,params)
        self.cnxn.commit()
        QMessageBox.information(self,'提示信息','数据采集完成,请检查!',QMessageBox.Yes)

    def Data_Query(self):
        date_start = self.DateEdit_Start.text()
        date_end = self.DateEdit_End.text()
        employeeid = "%" + self.LineEdit_ID.text() + "%"
        params = ([date_start,date_end,employeeid])
        sql = "select a.employeeid,b.name,c.deptnm,b.position,a.card_d,a.card_tm,a.card_ip from em_card_food a left join zwerp.dbo.employee b on a.employeeid = b.employeeid left join zwerp.dbo.dept c on b.deptid = c.deptid where a.card_d between ? and ? and a.employeeid like ?"
        self.result = self.cursor.execute(sql,params).fetchall()
        model = QStandardItemModel(len(self.result),7)
        model.setHorizontalHeaderLabels(['工号','姓名','组别','职称','就餐日期','就餐时间','卡机IP'])
        for row in range(len(self.result)):
            for column in range(7):
                item = QStandardItem(self.result[row][column])
                model.setItem(row,column,item)
        self.TableView_Data.setModel(model)
        self.statusBar().showMessage('查询成功，累计查询到' + str(len(self.result)) + '笔数据!')

    #数据分析,调用存储过程
    def Data_Ana(self):
        date_start = self.DateEdit_Start.text()
        date_end = self.DateEdit_End.text()
        params = (date_start,date_end)
        self.cursor.execute("{CALL Pro_generate_food_data (?,?)}",params)
        
    #导出EXCEL
    def Export_Data(self):
        date_start = self.DateEdit_Start.text()
        date_end = self.DateEdit_End.text()
        if date_start is None or len(date_start.strip()) == 0 or date_end is None or len(date_end.strip()) ==0:
            QMessageBox(self,'提示信息','导出起始日期不能为空!')
            return
        params = ([date_start,date_end])
        sql = "select a.employeeid,b.name,c.deptnm,b.position,a.card_d,a.zc_record,a.zc_times,a.wc_record,a.wc_times,a.food_times,a.food_record\
               from em_card_food_ana a left join zwerp.dbo.Employee b on a.employeeid = b.EmployeeID\
				left join zwerp.dbo.dept c on b.deptID = c.Deptid where card_d between ? and ? order by b.deptid,a.employeeid,a.card_d "
        result = self.cursor.execute(sql,params).fetchall()
        if len(result) == 0:
            QMessageBox.information(self,'提示信息','无资料可供导出',QMessageBox.Yes)
            return
        file_name = os.path.join(os.path.abspath('.'),date_start + "至" + date_end + "就餐记录分析.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = ('就餐数据分析')
        ws.append(['工号','姓名','组别','职称','就餐日期','中餐记录','中餐次数','晚餐记录','晚餐次数','累计次数','就餐记录'])
        for n in result:
            ws.append([n[0],n[1],n[2],n[3],n[4],n[5],n[6],n[7],n[8],n[9],n[10]])

        xfont = Font(size=14)
        xfill = PatternFill(fill_type=None,fgColor="0033CCCC")
        for column in [chr(n) for n in range(65,76)]:
            #设置字符格式
            ws[column + str(1)].font = Font(name='微软雅黑',size=14,color='00000000')
            #设置背景色
            ws[column + str(1)].fill = PatternFill("solid", fgColor="0099CCFF")
            #设置列宽
            ws.column_dimensions[column].width = 17
            #居中对齐
            ws[column + str(1)].alignment = Alignment(horizontal='center',vertical='center')

        #交替填充背景色      
        #for row in range(3,ws.max_row):
        #    if row%2 != 0:
        #        for column in [chr(n) for n in range(65,76)]:
        #            ws[column + str(row)].fill = PatternFill("solid", fgColor="0099CCFF")

        xborder = Border(left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
        for row in range(1,ws.max_row+1):
            for column in [chr(n) for n in range(65,76)]:
                ws[column + str(row)].font = Font(name='微软雅黑',size=13)
                ws[column + str(row)].border = xborder 

        wb.save(file_name)
        os.startfile(file_name)


if __name__ == "__main__":
    app = QApplication()
    win = Food_Analysis()
    win.show()
    app.exec_()